"""Build JSON datasets for the Efectividad Mecanicos web app from the Excel sources.

Outputs:
  - app/data/tareas.json     : detail rows from 'Efectividad Mecanicos.xlsx' (Hoja1, etc.)
  - app/data/ordenes.json    : {DocNum -> Tipo de orden} from 'Listado de Ordenes de Mantenimiento.xlsx'
  - app/data/repartido.json  : precomputed "Hs Mecanico Repartido" per (codigo_empleado, yyyy-mm)
                               replicating the DAX measure:
                                 Tiempo Std (unico y repartido) =
                                     VAR BaseUnicaPorMec = SUMMARIZE(Hoja1,
                                         Hoja1[DocNum], Hoja1[Fecha Inicio (dia)],
                                         Hoja1[Codigo tarea], Hoja1[Codigo empleado],
                                         "StdTarea", MAX(Hoja1[TiempoEstandarPorMecanico]))
                                     VAR ConReparto = ADDCOLUMNS(BaseUnicaPorMec,
                                         "StdAsignado", DIVIDE([StdTarea],
                                             DISTINCTCOUNT(Hoja1[Codigo empleado])
                                             ALLEXCEPT(Hoja1, DocNum, Fecha, Codigo tarea)))
                                     RETURN SUMX(ConReparto, [StdAsignado])

Run from the project folder whenever the Excel files change:
    python app/build_data.py
"""
import json
import datetime
from collections import defaultdict
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent
TAREAS_XLSX = BASE / "Efectividad Mecanicos.xlsx"
ORDENES_XLSX = BASE / "Listado de Ordenes de Mantenimiento.xlsx"
OUT = BASE / "app" / "data"
OUT.mkdir(parents=True, exist_ok=True)


def _to_minutes(v):
    """Coerce Excel cell to a number of minutes. Excel sometimes stores durations as time/datetime."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, datetime.time):
        return v.hour * 60 + v.minute + v.second / 60.0
    if isinstance(v, datetime.timedelta):
        return v.total_seconds() / 60.0
    if isinstance(v, datetime.datetime):
        return (v - datetime.datetime(1899, 12, 30)).total_seconds() / 60.0
    if isinstance(v, str):
        s = v.strip().replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return 0.0
    return 0.0


def _to_date(v):
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.datetime.strptime(s.split()[0], fmt).date().isoformat()
            except ValueError:
                pass
    return None


def load_tareas():
    """Read task rows from every relevant sheet of 'Efectividad Mecanicos.xlsx'."""
    wb = openpyxl.load_workbook(TAREAS_XLSX, data_only=True)
    rows_out = []
    sheet_map = {
        "Enero": 1,
        "Hoja1": None,
        "Hoja6": None,
        "Hoja2": None,
        "Hoja5": None,
        "Hoja1 (tdu enero y diciembre)": 1,
        "Hoja3": 1,
    }

    for sheet_name, _ in sheet_map.items():
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers) if h}
        doc_col = idx.get("DocNum")
        emp_cod = idx.get("Codigo empleado")
        emp_nom = idx.get("Empleado") or idx.get("Nombre empleado")
        fecha_col = idx.get("Fecha inicio")
        trab_col = idx.get("Horas Trabajadas (minutos)") or idx.get("Horas trabajadas")
        std_col = idx.get("TiempoEstandarPorMecanico")
        dif_col = idx.get("Dif HTrabajadas vs TExMecanico")
        tarea_col = idx.get("Tarea")
        tarea_cod_col = idx.get("Codigo tarea")

        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or all(v is None for v in r):
                continue
            fecha = _to_date(r[fecha_col]) if fecha_col is not None else None
            if fecha is None:
                continue
            rows_out.append({
                "docNum": r[doc_col] if doc_col is not None else None,
                "empleadoCodigo": str(r[emp_cod]) if emp_cod is not None and r[emp_cod] is not None else None,
                "empleadoNombre": r[emp_nom] if emp_nom is not None else None,
                "fecha": fecha,
                "horasTrabajadasMin": _to_minutes(r[trab_col]),
                "tiempoEstandarMin": _to_minutes(r[std_col]) if std_col is not None else 0,
                "difMin": _to_minutes(r[dif_col]) if dif_col is not None else 0,
                "tarea": r[tarea_col] if tarea_col is not None else None,
                "codigoTarea": r[tarea_cod_col] if tarea_cod_col is not None else None,
            })

    OUT.joinpath("tareas.json").write_text(json.dumps(rows_out, ensure_ascii=False), encoding="utf-8")
    return rows_out


def load_ordenes():
    """Read 'Listado de Ordenes de Mantenimiento.xlsx' and return a {DocNum -> Tipo} map."""
    wb = openpyxl.load_workbook(ORDENES_XLSX, data_only=True)
    ws = wb["Hoja1"]
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        nro = r[0]
        tipo = r[1]
        if nro is None:
            continue
        out[str(int(nro))] = tipo
    OUT.joinpath("ordenes.json").write_text(json.dumps(out, ensure_ascii=False), encoding="utf-8")
    return out


def compute_repartido(tareas):
    """Replicate the DAX measure 'Hs Mecanico Repartido' for backward compat."""
    metrics = compute_metrics(tareas)
    flat = {ym: {emp: v["hsTot"] for emp, v in d.items()} for ym, d in metrics.items()}
    OUT.joinpath("repartido.json").write_text(json.dumps(flat, ensure_ascii=False), encoding="utf-8")
    return flat


def compute_metrics(tareas):
    """Replicate the DAX measures from the PBI model.

    - 'Hs Totales' = 'Hs Mecanico Repartido':
        unique (DocNum, fecha, codigo_tarea, codigo_empleado), MAX(std) per group,
        divide by distinct employees on the same (DocNum, fecha, codigo_tarea),
        then SUMX.

    - 'Margen Productividad (Filtrado)' =
        DIVIDE([Suma Tiempo Estandar (No Negativos)],
               [Horas Reales (Optimas)]) - 1
       where Suma Tiempo Estandar (No Negativos) = SUM(std) on rows with dif>=0
             Horas Reales (Optimas) = SUM(horas trabajadas) on rows with dif>=0

    - 'Efectividad' (calculated column) = IF(dif >= 0, "Productivo", "No Productivo")

    - 'Contar Productivo' = CALCULATE(COUNTROWS(Hoja1), Hoja1[Efectividad] = "Productivo")

    - 'Productivo Total' = COUNT(Hoja1[Efectividad])

    - '% Procutivo' = [Contar Productivo] * 100 / [Productivo Total] / 100
                    = COUNT of rows where dif>=0 / COUNT of all rows

    - 'Superan las 80Hs' = IF(Hoja1[Hs Mecanico Repartido] >= 80, "SI", "NO")

    - 'Supera %50 Optimas' = IF(Hoja1[% Procutivo] >= 0.50, "SI", "NO")
    """
    # 1) Find distinct employees per (docNum, fecha, codigoTarea) for the repartido measure
    key_employees = defaultdict(set)
    for r in tareas:
        key_employees[(str(r["docNum"]), r["fecha"], r["codigoTarea"])].add(r["empleadoCodigo"])

    # 2) Precompute per-row (docNum, fecha, codigoTarea, codigoEmpleado) -> MAX(std)
    seen_keys = {}
    for r in tareas:
        key = (str(r["docNum"]), r["fecha"], r["codigoTarea"], r["empleadoCodigo"])
        if r["tiempoEstandarMin"] > seen_keys.get(key, 0):
            seen_keys[key] = r["tiempoEstandarMin"]

    # 3) Per (empleado, yyyy-mm): HsTot, HsMec, HsMec(pos), HsStd(pos), total_rows, prod_rows
    per = defaultdict(lambda: {
        "hsTot": 0.0, "trab": 0.0, "trab_pos": 0.0, "std_pos": 0.0,
        "rows_total": 0, "rows_prod": 0,
    })

    for (doc, fecha, tarea, emp), max_std in seen_keys.items():
        cant = len(key_employees[(doc, fecha, tarea)])
        if cant > 0:
            std_asignado = max_std / cant
            ym = fecha[:7]
            per[(emp, ym)]["hsTot"] += std_asignado

    for r in tareas:
        k = (r["empleadoCodigo"], r["fecha"][:7])
        per[k]["trab"] += r["horasTrabajadasMin"]
        per[k]["rows_total"] += 1
        if r["difMin"] >= 0:
            per[k]["trab_pos"] += r["horasTrabajadasMin"]
            per[k]["std_pos"] += r["tiempoEstandarMin"]
            per[k]["rows_prod"] += 1

    out = defaultdict(dict)
    for (emp, ym), v in per.items():
        hsTot = v["hsTot"] / 60.0
        hsMec = v["trab"] / 60.0
        trab_pos = v["trab_pos"] / 60.0
        std_pos = v["std_pos"] / 60.0
        # Margen Productividad (Filtrado) = DIVIDE(StdPos, TrabPos) - 1
        margen = (std_pos / trab_pos - 1) * 100 if trab_pos else 0
        # % Procutivo = Contar Productivo / Productivo Total * 100
        pctProd = (v["rows_prod"] / v["rows_total"] * 100) if v["rows_total"] else 0
        # Superan las 80Hs = IF(HsTot >= 80, "SI", "NO")
        supera80 = "SI" if hsTot >= 80 else "NO"
        # Supera %50 Optimas = IF(%Procutivo >= 50, "SI", "NO")
        supera50 = "SI" if pctProd >= 50 else "NO"
        out[ym][emp] = {
            "hsTot": round(hsTot, 4),
            "hsMec": round(hsMec, 4),
            "margen": round(margen, 4),
            "pctProd": round(pctProd, 4),
            "supera80": supera80,
            "supera50": supera50,
        }

    # Per-month totals (PBI's "Total" row recomputes the formulas across all rows)
    per_ym = defaultdict(lambda: {
        "hsTot": 0.0, "trab": 0.0, "trab_pos": 0.0, "std_pos": 0.0,
        "rows_total": 0, "rows_prod": 0,
    })
    for (doc, fecha, tarea, emp), max_std in seen_keys.items():
        cant = len(key_employees[(doc, fecha, tarea)])
        if cant > 0:
            ym = fecha[:7]
            per_ym[ym]["hsTot"] += max_std / cant
    for r in tareas:
        ym = r["fecha"][:7]
        per_ym[ym]["trab"] += r["horasTrabajadasMin"]
        per_ym[ym]["rows_total"] += 1
        if r["difMin"] >= 0:
            per_ym[ym]["trab_pos"] += r["horasTrabajadasMin"]
            per_ym[ym]["std_pos"] += r["tiempoEstandarMin"]
            per_ym[ym]["rows_prod"] += 1

    for ym, v in per_ym.items():
        hsTot = v["hsTot"] / 60.0
        hsMec = v["trab"] / 60.0
        trab_pos = v["trab_pos"] / 60.0
        std_pos = v["std_pos"] / 60.0
        margen = (std_pos / trab_pos - 1) * 100 if trab_pos else 0
        pctProd = (v["rows_prod"] / v["rows_total"] * 100) if v["rows_total"] else 0
        out[ym]["__total__"] = {
            "hsTot": round(hsTot, 4),
            "hsMec": round(hsMec, 4),
            "margen": round(margen, 4),
            "pctProd": round(pctProd, 4),
            "supera80": "SI" if hsTot >= 80 else "NO",
            "supera50": "SI" if pctProd >= 50 else "NO",
            "isTotal": True,
        }

    OUT.joinpath("metricas.json").write_text(json.dumps(out, ensure_ascii=False), encoding="utf-8")
    return out


if __name__ == "__main__":
    print("Loading tareas...")
    t = load_tareas()
    print(f"  -> {len(t):,} rows")
    print("Loading ordenes...")
    o = load_ordenes()
    print(f"  -> {len(o):,} orders")
    print("Computing PBI metrics...")
    r = compute_metrics(t)
    print(f"  -> {sum(len(v) for v in r.values()):,} (empleado, mes) pairs")
    print("Done.")
