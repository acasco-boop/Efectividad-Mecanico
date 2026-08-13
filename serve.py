"""Servidor HTTP dinámico para Efectividad Mecanicos.

Lee los archivos Excel en app/excel/ y re-calcula las metricas cada vez que el
archivo es modificado (deteccion por mtime). Sirve la web estatica + los datos
JSON en el endpoint /api/data.

Uso:
    python app/serve.py
    python app/serve.py 8765

Requisitos: pip install openpyxl
"""
import datetime
import http.server
import json
import os
import socketserver
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

APP_DIR = Path(__file__).resolve().parent
EXCEL_DIR = APP_DIR / "excel"
TAREAS_XLSX = EXCEL_DIR / "efectividad.xlsx"
ORDENES_XLSX = EXCEL_DIR / "ordenes.xlsx"

PORT = int(sys.argv[1]) if len(sys.argv) > 1 else 8765

CACHE = {}  # path -> (mtime, payload)


# ---------------------------------------------------------------------------
# Lectura y calculo (equivalente a build_data.py, pero en memoria)
# ---------------------------------------------------------------------------
def _to_minutes(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, datetime.time):
        return v.hour * 60 + v.minute + v.second / 60.0
    if isinstance(v, datetime.timedelta):
        return v.total_seconds() / 60.0
    if isinstance(v, datetime.datetime):
        return (v - datetime.datetime(1899, 12, 30)).total_seconds() / 60.0
    if isinstance(v, str):
        s = v.strip().replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return 0.0
    return 0.0


def _to_date(v):
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.datetime.strptime(s.split()[0], fmt).date().isoformat()
            except ValueError:
                pass
    return None


def load_tareas():
    wb = openpyxl.load_workbook(str(TAREAS_XLSX), data_only=True)
    rows_out = []
    sheet_map = {
        "Enero": 1,
        "Hoja1": None,
        "Hoja6": None,
        "Hoja2": None,
        "Hoja5": None,
        "Hoja1 (tdu enero y diciembre)": 1,
        "Hoja3": 1,
    }
    for sheet_name in sheet_map:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers) if h}
        doc_col = idx.get("DocNum")
        emp_cod = idx.get("Codigo empleado")
        emp_nom = idx.get("Empleado") or idx.get("Nombre empleado")
        fecha_col = idx.get("Fecha inicio")
        trab_col = idx.get("Horas Trabajadas (minutos)") or idx.get("Horas trabajadas")
        std_col = idx.get("TiempoEstandarPorMecanico")
        dif_col = idx.get("Dif HTrabajadas vs TExMecanico")
        tarea_col = idx.get("Tarea")
        tarea_cod_col = idx.get("Codigo tarea")
        if fecha_col is None:
            continue
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or all(v is None for v in r):
                continue
            fecha = _to_date(r[fecha_col])
            if fecha is None:
                continue
            rows_out.append({
                "docNum": r[doc_col] if doc_col is not None else None,
                "empleadoCodigo": str(r[emp_cod]) if emp_cod is not None and r[emp_cod] is not None else None,
                "empleadoNombre": r[emp_nom] if emp_nom is not None else None,
                "fecha": fecha,
                "horasTrabajadasMin": _to_minutes(r[trab_col]) if trab_col is not None else 0.0,
                "tiempoEstandarMin": _to_minutes(r[std_col]) if std_col is not None else 0.0,
                "difMin": _to_minutes(r[dif_col]) if dif_col is not None else 0.0,
                "tarea": r[tarea_col] if tarea_col is not None else None,
                "codigoTarea": r[tarea_cod_col] if tarea_cod_col is not None else None,
            })
    return rows_out


def load_ordenes():
    wb = openpyxl.load_workbook(str(ORDENES_XLSX), data_only=True)
    ws = wb["Hoja1"]
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        out[str(int(r[0]))] = r[1]
    return out


def compute_metrics(tareas):
    key_employees = defaultdict(set)
    for r in tareas:
        key_employees[(str(r["docNum"]), r["fecha"], r["codigoTarea"])].add(r["empleadoCodigo"])

    seen_keys = {}
    for r in tareas:
        key = (str(r["docNum"]), r["fecha"], r["codigoTarea"], r["empleadoCodigo"])
        if r["tiempoEstandarMin"] > seen_keys.get(key, 0):
            seen_keys[key] = r["tiempoEstandarMin"]

    per = defaultdict(lambda: {"hsTot": 0.0, "trab": 0.0, "trab_pos": 0.0,
                               "std_pos": 0.0, "rows_total": 0, "rows_prod": 0})
    for (doc, fecha, tarea, emp), max_std in seen_keys.items():
        cant = len(key_employees[(doc, fecha, tarea)])
        if cant > 0:
            per[(emp, fecha[:7])]["hsTot"] += max_std / cant
    for r in tareas:
        k = (r["empleadoCodigo"], r["fecha"][:7])
        per[k]["trab"] += r["horasTrabajadasMin"]
        per[k]["rows_total"] += 1
        if r["difMin"] >= 0:
            per[k]["trab_pos"] += r["horasTrabajadasMin"]
            per[k]["std_pos"] += r["tiempoEstandarMin"]
            per[k]["rows_prod"] += 1

    out = defaultdict(dict)
    for (emp, ym), v in per.items():
        hsTot = v["hsTot"] / 60.0
        hsMec = v["trab"] / 60.0
        trab_pos = v["trab_pos"] / 60.0
        std_pos = v["std_pos"] / 60.0
        margen = (std_pos / trab_pos - 1) * 100 if trab_pos else 0
        pctProd = (v["rows_prod"] / v["rows_total"] * 100) if v["rows_total"] else 0
        out[ym][emp] = {
            "hsTot": round(hsTot, 4),
            "hsMec": round(hsMec, 4),
            "margen": round(margen, 4),
            "pctProd": round(pctProd, 4),
            "supera80": "SI" if hsTot >= 80 else "NO",
            "supera50": "SI" if pctProd >= 50 else "NO",
        }

    per_ym = defaultdict(lambda: {"hsTot": 0.0, "trab": 0.0, "trab_pos": 0.0,
                                 "std_pos": 0.0, "rows_total": 0, "rows_prod": 0})
    for (doc, fecha, tarea, emp), max_std in seen_keys.items():
        cant = len(key_employees[(doc, fecha, tarea)])
        if cant > 0:
            per_ym[fecha[:7]]["hsTot"] += max_std / cant
    for r in tareas:
        ym = r["fecha"][:7]
        per_ym[ym]["trab"] += r["horasTrabajadasMin"]
        per_ym[ym]["rows_total"] += 1
        if r["difMin"] >= 0:
            per_ym[ym]["trab_pos"] += r["horasTrabajadasMin"]
            per_ym[ym]["std_pos"] += r["tiempoEstandarMin"]
            per_ym[ym]["rows_prod"] += 1

    for ym, v in per_ym.items():
        hsTot = v["hsTot"] / 60.0
        hsMec = v["trab"] / 60.0
        trab_pos = v["trab_pos"] / 60.0
        std_pos = v["std_pos"] / 60.0
        margen = (std_pos / trab_pos - 1) * 100 if trab_pos else 0
        pctProd = (v["rows_prod"] / v["rows_total"] * 100) if v["rows_total"] else 0
        out[ym]["__total__"] = {
            "hsTot": round(hsTot, 4),
            "hsMec": round(hsMec, 4),
            "margen": round(margen, 4),
            "pctProd": round(pctProd, 4),
            "supera80": "SI" if hsTot >= 80 else "NO",
            "supera50": "SI" if pctProd >= 50 else "NO",
            "isTotal": True,
        }
    return dict(out)


def compute_all():
    """Lee los Exceles (cacheados por mtime) y devuelve los 3 datasets."""
    out = {}
    for key, xlsx, fn in (
        ("tareas", TAREAS_XLSX, load_tareas),
        ("ordenes", ORDENES_XLSX, load_ordenes),
    ):
        mtime = xlsx.stat().st_mtime
        cached = CACHE.get(xlsx)
        if cached and cached[0] == mtime:
            out[key] = cached[1]
        else:
            data = fn()
            CACHE[xlsx] = (mtime, data)
            out[key] = data
            print(f"  [reload] {xlsx.name} -> {len(data):,} rows (mtime={datetime.datetime.fromtimestamp(mtime):%H:%M:%S})")
    # metricas depende de tareas
    tareas_mtime = TAREAS_XLSX.stat().st_mtime
    metricas_cached = CACHE.get("__metricas__")
    if metricas_cached and metricas_cached[0] == tareas_mtime:
        out["metricas"] = metricas_cached[1]
    else:
        m = compute_metrics(out["tareas"])
        CACHE["__metricas__"] = (tareas_mtime, m)
        out["metricas"] = m
        total = sum(len(v) for v in m.values())
        print(f"  [reload] metricas -> {total:,} (empleado, mes) pairs")
    return out


# ---------------------------------------------------------------------------
# Servidor HTTP
# ---------------------------------------------------------------------------
class Handler(http.server.BaseHTTPRequestHandler):
    def do_GET(self):
        path = self.path.split("?", 1)[0]
        if path in ("/", "/index.html"):
            return self._serve_file(APP_DIR / "index.html", "text/html")
        if path == "/api/data":
            return self._serve_api()
        # Otros assets estaticos: servir desde APP_DIR (css/js/img/etc.)
        if path.startswith("/"):
            full = (APP_DIR / path.lstrip("/")).resolve()
            if APP_DIR.resolve() in full.parents and full.exists() and full.is_file():
                ctype = "text/plain"
                lower = full.name.lower()
                if lower.endswith(".js"): ctype = "application/javascript"
                elif lower.endswith(".css"): ctype = "text/css"
                elif lower.endswith(".html"): ctype = "text/html"
                elif lower.endswith(".json"): ctype = "application/json"
                elif lower.endswith(".svg"): ctype = "image/svg+xml"
                elif lower.endswith(".png"): ctype = "image/png"
                return self._serve_file(full, ctype)
        self.send_error(404, f"Not found: {path}")

    def _serve_file(self, full_path, content_type):
        body = full_path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-cache")
        self.end_headers()
        self.wfile.write(body)

    def _serve_api(self):
        try:
            data = compute_all()
        except FileNotFoundError as e:
            self.send_error(503, f"Excel no encontrado: {e.filename}")
            return
        except Exception as e:
            self.send_error(500, f"Error procesando Excel: {e}")
            return
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, fmt, *args):
        sys.stderr.write(f"[{datetime.datetime.now():%H:%M:%S}] {self.address_string()} {fmt%args}\n")


class ReusableTCPServer(socketserver.ThreadingTCPServer):
    allow_reuse_address = True
    daemon_threads = True


def main():
    print(f"Efectividad Mecanicos server")
    print(f"  Excel dir:   {EXCEL_DIR}")
    print(f"  Tareas:      {TAREAS_XLSX.name} ({'OK' if TAREAS_XLSX.exists() else 'FALTA'})")
    print(f"  Ordenes:     {ORDENES_XLSX.name} ({'OK' if ORDENES_XLSX.exists() else 'FALTA'})")
    print(f"  Listening:   http://0.0.0.0:{PORT}")
    print(f"  Deteniendo con Ctrl+C.")
    srv = ReusableTCPServer(("0.0.0.0", PORT), Handler)
    try:
        srv.serve_forever()
    except KeyboardInterrupt:
        print("\nApagando servidor...")
    finally:
        srv.server_close()


if __name__ == "__main__":
    main()
